import os
import openpyxl
import zipfile

def find_xlsx_file(folder):
    """递归查找文件夹中的.xlsx文件，排除临时文件"""
    for root, dirs, files in os.walk(folder):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~$'):  # 排除 Excel 临时文件
                return os.path.join(root, file)
    return None

def is_valid_xlsx(file_path):
    """检查文件是否是有效的.xlsx文件（ZIP格式）"""
    try:
        with zipfile.ZipFile(file_path, 'r') as zf:
            zf.testzip()  # 测试 ZIP 文件完整性
        return True
    except zipfile.BadZipFile:
        return False

def extract_data(xlsx_path):
    """从xlsx文件中提取姓名、学号、C1-C3和D1-D6，获取公式计算后的值"""
    try:
        if not is_valid_xlsx(xlsx_path):
            raise ValueError(f"{xlsx_path} 不是有效的 .xlsx 文件")
        
        # 使用 data_only=True 获取公式计算后的值
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        sheet = wb.active
        
        # 提取姓名和学号
        name = sheet['E2'].value
        student_id = sheet['I2'].value or sheet['J2'].value
        
        # 查找C1-C3（L列，第12列）
        c1, c2, c3 = 0, 0, 0
        for row in sheet.iter_rows(min_col=12, max_col=12):
            cell_value = row[0].value
            if cell_value == "C1总分":
                c1 = sheet.cell(row=row[0].row + 1, column=12).value or 0
            elif cell_value == "C2总分":
                c2 = sheet.cell(row=row[0].row + 1, column=12).value or 0
            elif cell_value == "C3总分":
                c3 = sheet.cell(row=row[0].row + 1, column=12).value or 0
        
        # 查找D1-D6（K列，第11列）
        d1, d2, d3, d4, d5, d6 = 0, 0, 0, 0, 0, 0
        for row in sheet.iter_rows(min_col=11, max_col=11):
            cell_value = row[0].value
            if cell_value == "D1总分":
                d1 = sheet.cell(row=row[0].row + 1, column=11).value or 0
            elif cell_value == "D2总分":
                d2 = sheet.cell(row=row[0].row + 1, column=11).value or 0
            elif cell_value == "D3总分":
                d3 = sheet.cell(row=row[0].row + 1, column=11).value or 0
            elif cell_value == "D4总分":
                d4 = sheet.cell(row=row[0].row + 1, column=11).value or 0
            elif cell_value == "D5总分":
                d5 = sheet.cell(row=row[0].row + 1, column=11).value or 0
            elif cell_value == "D6总分":
                d6 = sheet.cell(row=row[0].row + 1, column=11).value or 0
        
        return name, student_id, c1, c2, c3, d1, d2, d3, d4, d5, d6
    
    except Exception as e:
        print(f"提取数据失败 {xlsx_path}: {e}")
        return None

def main():
    # 找到综测成绩汇总表
    try:
        summary_file = [f for f in os.listdir() if '综测成绩汇总表' in f and f.endswith('.xlsx')][0]
    except IndexError:
        print("错误：未找到包含‘综测成绩汇总表’的 .xlsx 文件")
        return
    
    if not is_valid_xlsx(summary_file):
        print(f"错误：{summary_file} 不是有效的 .xlsx 文件")
        return
    
    summary_wb = openpyxl.load_workbook(summary_file, data_only=True)  # 确保汇总表也读取公式值
    summary_sheet = summary_wb.active
    
    # 清空现有数据（从第二行开始），仅清空 C1, C2, C3, D1-D6 列
    for row in summary_sheet.iter_rows(min_row=2):
        for col in [5, 6, 7, 9, 10, 11, 12, 13, 14]:  # C1=5, C2=6, C3=7, D1=9, ..., D6=14
            row[col - 1].value = None  # 清空指定列
    
    # 从第二行开始写入新数据
    row = 2
    success_count = 0
    total_count = 0
    
    # 遍历所有文件夹，排除 env 文件夹
    for folder in os.listdir():
        if os.path.isdir(folder) and folder != 'env':
            total_count += 1
            xlsx_file = find_xlsx_file(folder)
            if xlsx_file:
                print(f"处理文件夹 {folder} 中的文件: {xlsx_file}")
                data = extract_data(xlsx_file)
                if data:
                    name, student_id, c1, c2, c3, d1, d2, d3, d4, d5, d6 = data
                    # 写入汇总表
                    # 表头：姓名 A B C1 C2 C3 C D1 D2 D3 D4 D5 D6
                    # 计算C和D
                    c = c1 + c2 + c3
                    d = d1 + d2 + d3 + d4 + d5 - d6
                    
                    # A和B暂设为0（可根据需求调整）
                    a = 0
                    b = 0
                    s = a + b * 0.85 + c * 0.10 + d * 0.05
                    summary_sheet.cell(row=row, column=2).value = name      # 姓名 (第1列)
                    # A 和 B 不填写，保持为空 (第2、3列)
                    summary_sheet.cell(row=row, column=5).value = c1        # C1 (第4列)
                    summary_sheet.cell(row=row, column=6).value = c2        # C2 (第5列)
                    summary_sheet.cell(row=row, column=7).value = c3        # C3 (第6列)
                    summary_sheet.cell(row=row, column=8).value = c
                    # C 不填写 (第7列)
                    summary_sheet.cell(row=row, column=9).value = d1       # D1 (第8列)
                    summary_sheet.cell(row=row, column=10).value = d2       # D2 (第9列)
                    summary_sheet.cell(row=row, column=11).value = d3      # D3 (第10列)
                    summary_sheet.cell(row=row, column=12).value = d4      # D4 (第11列)
                    summary_sheet.cell(row=row, column=13).value = d5      # D5 (第12列)
                    summary_sheet.cell(row=row, column=14).value = d6      # D6 (第13列)
                    summary_sheet.cell(row=row, column=15).value = d        # D
                    summary_sheet.cell(row=row, column=16).value = s        # S
                    row += 1
                    success_count += 1
                else:
                    print(f"跳过 {folder}，尝试处理下一个文件夹")
            else:
                print(f"未在文件夹 {folder} 中找到有效的 .xlsx 文件")
    
    # 保存汇总表
    summary_wb.save(summary_file)
    print(f"数据已整理并保存至 {summary_file}，成功处理 {success_count}/{total_count} 个文件夹")

if __name__ == "__main__":
    main()